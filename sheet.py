import openpyxl
from openpyxl.styles.borders import Border, Side
from openpyxl.styles import Alignment, PatternFill, Color

THIN_BORDER = Border(left=Side(style='thin'),
                     right=Side(style='thin'),
                     top=Side(style='thin'),
                     bottom=Side(style='thin'))

CENTER_ALIGNMENT = Alignment(horizontal='center', vertical='center')

FILL_COLOR = PatternFill(fgColor=Color('E5DE00'), fill_type='solid')
KEY_ELEMENT_FILL_COLOR = PatternFill(
    fgColor=Color('D30000'), fill_type='solid')


def set_cell_value(cell: openpyxl.cell, value):
    cell.alignment = CENTER_ALIGNMENT
    cell.border = THIN_BORDER
    cell.value = value


def append_row(ws: openpyxl.Workbook, start_pos: tuple[int, int], row: list) -> None:
    row_pos, col_pos = start_pos
    for index, value in enumerate(row):
        set_cell_value(
            ws.cell(row=row_pos, column=col_pos + index), str(value))


def color_key_element_cell(cell: openpyxl):
    cell.fill = KEY_ELEMENT_FILL_COLOR


def color_cell(cell: openpyxl.cell, color):
    cell.fill = color


def color_range(ws: openpyxl.Workbook, start_pos: tuple[int, int], end_pos: tuple[int, int]):
    start_row, start_col = start_pos
    end_row, end_col = end_pos

    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            color_cell(ws.cell(row=row, column=col), FILL_COLOR)


def talbe_start_position(coefficients, table_row, table_col):
    return table_row - 4 - len(coefficients), table_col


def get_leading_row_range_in_sheet(coefficients, leading_row, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS):
    start_row, start_col = talbe_start_position(
        coefficients, table_row, table_col)
    return (start_row + leading_row + NUMBER_OF_ADDITIONAL_ROWS, start_col), (start_row + leading_row + NUMBER_OF_ADDITIONAL_ROWS, start_col + len(coefficients[0]) + NUMBER_OF_ADDITIONAL_COLS - 1)


def get_leading_column_range_in_sheet(coefficients, leading_col, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS):
    start_row, start_col = talbe_start_position(
        coefficients, table_row, table_col)
    return (start_row, start_col + leading_col + NUMBER_OF_ADDITIONAL_COLS), (start_row + len(coefficients) + 1, start_col + leading_col + NUMBER_OF_ADDITIONAL_COLS)


def get_key_element_position(coefficients, leading_row, leading_column, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS):
    start_row, start_col = talbe_start_position(
        coefficients, table_row, table_col)
    return start_row + leading_row + NUMBER_OF_ADDITIONAL_ROWS, start_col + leading_column + NUMBER_OF_ADDITIONAL_COLS


def colorize_table(ws, coefficients, leading_row, leading_column, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS):
    color_range(
        ws, *get_leading_column_range_in_sheet(coefficients, leading_column, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS))
    color_range(
        ws, *get_leading_row_range_in_sheet(coefficients, leading_row, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS))

    key_element_position_in_sheet = get_key_element_position(
        coefficients, leading_row, leading_column, table_row, table_col, NUMBER_OF_ADDITIONAL_ROWS, NUMBER_OF_ADDITIONAL_COLS)
    color_key_element_cell(ws.cell(
        row=key_element_position_in_sheet[0], column=key_element_position_in_sheet[1]))
